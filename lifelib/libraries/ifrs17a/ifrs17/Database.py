import dataclasses
from dataclasses import dataclass
import uuid
from typing import (Iterable, Optional, Union, Callable, Any)
import inspect
from collections import namedtuple

import pandas as pd
import openpyxl as opyxl

from .DataStructure import *
from .BaseClasses import BaseDatabase
from .Validations import *

__all__ = ('IfrsDatabase',)

# Current and Previous Parameters

TempName1 = namedtuple('TempName1', ['key', 'currentPeriod', 'previousPeriod'])


@dataclass
class IGroupingParams:
    Key: Any
    Values: list[Any]


def remove_duplicates(input_list: list) -> list:
    seen = set()
    return [item for item in input_list if not (item in seen or seen.add(item))]


def replace_nan(df):
    float_cols = df.select_dtypes(include=['float64']).columns
    str_cols = df.select_dtypes(include=['object']).columns

    df.loc[:, float_cols] = df.loc[:, float_cols].fillna(0)
    df.loc[:, str_cols] = df.loc[:, str_cols].fillna('')
    return df


class IfrsDatabase(BaseDatabase):

    _import_formats: dict[ImportFormats, Callable] = {}

    @classmethod
    def DefineFormat(cls, format: ImportFormats, body: Callable[['IfrsDatabase', IDataSet], ImportArgs]):
        cls._import_formats[format] = body

    def ImportFile(self, pathToFile: str, format_: ImportFormats = '',
                   type_: Optional[Union[type, list[type]]] = None) -> ImportArgs:
        if type_:
            result = self._FromFileWithType(pathToFile, type_)
            for k, v in result.items():
                self.Update(k, v)
        else:
            sheets: dict[str, pd.DataFrame] = {}
            for name in opyxl.load_workbook(pathToFile).sheetnames:
                sheets[name] = df = replace_nan(pd.read_excel(pathToFile, sheet_name=name))

            return self._import_formats[format_](self, IDataSet(sheets))

    @staticmethod
    def _FromFileWithType(pathToFile: str,
                          type_: Optional[Union[type, list[type]]] = None):

        types_ = type_ if isinstance(type_, Iterable) else [type_]

        wb = opyxl.load_workbook(pathToFile)
        result = {}
        for tp in types_:
            rows = list(wb[tp.__name__].values)
            objs = []
            for r in rows[1:]:
                kwargs = {}
                extids = {}  # index: value
                vals = {}  # index: value
                for i, name in enumerate(rows[0]):
                    if name[:10] == "ExternalId":
                        extids[int(name[10:])] = r[i] if r[i] else ''
                    elif name[:6] == "Values":
                        vals[int(name[6:])] = float(r[i])
                    elif name == "InputSource":
                        kwargs[name] = [s.strip() for s in r[i].split(",")]
                    else:
                        kwargs[name] = r[i]

                if "ExternalId" in inspect.signature(tp).parameters.keys():
                    kwargs["ExternalId"] = list(extids[i] for i in extids.keys())
                elif "Values" in inspect.signature(tp).parameters.keys():
                    kwargs["Values"] = list(vals[i] for i in vals.keys())

                fields = set(f.name for f in dataclasses.fields(tp))
                args = set(kwargs)
                missing = fields - args

                if missing:
                    if 'Id' in missing:
                        kwargs['Id'] = uuid.uuid4()
                    if 'Name' in missing:
                        kwargs['Name'] = ''
                    if 'Scenario' in missing:
                        kwargs['Scenario'] = ''

                objs.append(tp(**kwargs))

            result[tp] = objs

        return result

    def FromDataSet(self, dataSet: IDataSet, type_: type, body: Callable = None, format_=None):

        if body:
            if format_:
                name = format_
            else:
                name = type_.__name__
            result = []
            for r in dataSet.Tables[name].to_dict('records'):
                result.append(body(dataSet, r))
            self.Update(type_, result)
        else:
            self.Update(type_, self._df_to_records(dataSet.Tables[type_.__name__], type_))

    @staticmethod
    def _df_to_records(df: pd.DataFrame, type_: type) -> list:

        records = []
        try:
            for r in df.to_dict('records'):
                records.append(type_(**r))

        except TypeError:
            records.clear()
            for r in df.to_dict('records'):
                r['Id'] = uuid.uuid4()
                records.append(type_(**r))

        return records

    def LoadCurrentParameter(
        self,
        type_,
        args: Args,
        identityExpression: Callable[[IWithYearAndMonth], str],
        filterExpression: Callable[[IWithYearAndMonth], bool] = None) -> dict[str, IWithYearAndMonth]:

        temp = [x for x in self.LoadParameter(type_, args.Year, args.Month, filterExpression)
                 if x.Scenario == args.Scenario or x.Scenario == '']
        groupby = {}
        for x in temp:
            k = identityExpression(x)
            groupby.setdefault(k, []).append(x)

        return {k: max(v, key=lambda y: y.Year * 100 + y.Month, default=None)for k, v in groupby.items()}   # Scenario not considered

    def LoadCurrentAndPreviousParameter(
        self,
        type_: type,
        args: Args,
        identityExpression: Callable[[IWithYearMonthAndScenario], str],
        filterExpression: Optional[Callable[[IWithYearMonthAndScenario], bool]] = None) -> dict[str, dict[int, IWithYearMonthAndScenario]]:

        parameters = {}
        for x in [yc for yc in self.LoadParameter(type_, args.Year, args.Month, filterExpression)
                  if yc.Scenario == args.Scenario or yc.Scenario == '']:

            k = identityExpression(x)
            parameters.setdefault(k, []).append(x)

        ret: dict[str, dict[int, IWithYearMonthAndScenario]] = {}

        for k, p in parameters.items():

            inner = ret.setdefault(k, {})
            currentCandidate = max([x for x in p if x.Year == args.Year], default=None, key= lambda y: y.Year * 100 + y.Month)
            previousCandidate = max([x for x in p if x.Year < args.Year and not x.Scenario], default=None, key= lambda y: y.Year * 100 + y.Month)
            currentCandidateBE = max([x for x in p if x.Year <= args.Year and not x.Scenario], default=None, key= lambda y: y.Year * 100 + y.Month)

            inner[CurrentPeriod] = currentCandidate if currentCandidate else previousCandidate
            inner[PreviousPeriod] = previousCandidate if previousCandidate else currentCandidateBE if currentCandidateBE else currentCandidate
            # TODO: log error if currentCandidate is null

        return ret

    # Yield Curve
    def LoadLockedInYieldCurve(self, args: Args, dataNodes: list[DataNodeData]) -> dict[str, YieldCurve]:

        lockedInYieldCurveByGoc: dict[str, YieldCurve] = dict()

        for dn in [x for x in dataNodes if x.ValuationApproach == ValuationApproaches.BBA]:

            argsNew = dataclasses.replace(args, Year=dn.Year, Month=dn.Month, Scenario=dn.Scenario)

            loadedYc = self.LoadCurrentParameter(YieldCurve, argsNew, lambda x: x.Currency, lambda x: x.Currency == dn.ContractualCurrency)

            if not (lockedYc := loadedYc.get(dn.ContractualCurrency, None)):
                raise YieldCurveNotFound

            lockedInYieldCurveByGoc[dn.DataNode] = lockedYc

        return lockedInYieldCurveByGoc

    def LoadCurrentYieldCurve(self,  args: Args, dataNodes: list[DataNodeData]) -> dict[str, dict[int, YieldCurve]]:
        contractualCurrenciesInScope = {dn.ContractualCurrency for dn in dataNodes}
        return self.LoadCurrentAndPreviousParameter(YieldCurve, args,
                lambda x: x.Currency,
                lambda x: x.Currency in contractualCurrenciesInScope)

    # Data Nodes
    def LoadDataNodes(self, args: Args) -> dict[str, DataNodeData]:

        dataNodeStates = self.LoadCurrentAndPreviousParameter(DataNodeState, args, lambda x: x.DataNode)
        activeDataNodes = [k for k, v in dataNodeStates.items() if v[CurrentPeriod].State != State.Inactive]

        temp = [dn for dn in self.Query(GroupOfContract) if dn.SystemName in activeDataNodes]
        result = {}
        for dn in temp:
            dnCurrentState = dataNodeStates[dn.SystemName][CurrentPeriod]
            dnPreviousState = dataNodeStates[dn.SystemName][PreviousPeriod]
            result[dn.SystemName] = DataNodeData(Year=dnPreviousState.Year,
                                      Month=dnPreviousState.Month,
                                      State=dnCurrentState.State,
                                      PreviousState=dnPreviousState.State,
                                      DataNode=dn.SystemName,
                                      ContractualCurrency=dn.ContractualCurrency,
                                      FunctionalCurrency=dn.FunctionalCurrency,
                                      LineOfBusiness=dn.LineOfBusiness,
                                      ValuationApproach=dn.ValuationApproach,
                                      OciType=dn.OciType,
                                      Portfolio=dn.Portfolio,
                                      AnnualCohort=dn.AnnualCohort,
                                      LiabilityType=dn.LiabilityType,
                                      Profitability=dn.Profitability,
                                      Partner=dn.Partner,
                                      IsReinsurance=isinstance(dn, GroupOfReinsuranceContract),
                                      Scenario='')

        return result

    def LoadParameter(
        self,
        type_: type,
        year: int,
        month: int,
        filterExpression: Optional[Callable[[IWithYearAndMonth], bool]] = None) -> list[IWithYearAndMonth]:

        result = [x for x in self.Query(type_) if x.Year == year and x.Month <= month or x.Year < year]

        if filterExpression:
            result = [x for x in result if filterExpression(x)]

        return result

    def LoadSingleDataNodeParameters(self, args: Args) -> dict[str, dict[int, SingleDataNodeParameter]]:
        return self.LoadCurrentAndPreviousParameter(SingleDataNodeParameter, args, lambda x: x.DataNode)

    def LoadInterDataNodeParameters(self, args: Args) -> dict[str, dict[int, set[InterDataNodeParameter]]]:

        identityExpressions: list[Callable[[InterDataNodeParameter], str]] = [lambda x: x.DataNode, lambda x: x.LinkedDataNode]
        parameterArray = self.LoadParameter(InterDataNodeParameter, args.Year, args.Month)

        parameters = []
        for ie in identityExpressions:
            temp = {}
            for p in parameterArray:
                k = ie(p)
                if k in temp:
                    temp[k].append(p)
                else:
                    temp[k] = [p]
            parameters.extend([IGroupingParams(Key=key, Values=val) for key, val in temp.items()])

        def _inner(p, gg):
            currentCandidate = max([x for x in gg if x.Year == args.Year], key=lambda x: x.Month, default=None) #[-1]
            previousCandidate = max([x for x in gg if x.Year < args.Year], key=lambda x: x.Year * 100 + x.Month, default=None)  #[-1]
            return TempName1(key=p.Key,
                 currentPeriod= currentCandidate if currentCandidate else previousCandidate,
                 previousPeriod= previousCandidate if previousCandidate else currentCandidate)

        result: list[TempName1] = []
        for p in parameters:
            groups = {}
            for x in p.Values:
                key = x.DataNode if x.DataNode != p.Key else x.LinkedDataNode
                if key in groups:
                    groups[key].append(x)
                else:
                    groups[key] = [x]

            for gg in groups.values():
                result.append(_inner(p, gg))


        temp = {}
        for x in result:
            if x.key in temp:
                temp[x.key].append(x)
            else:
                temp[x.key] = [x]

        temp2: dict[str, dict[int, set[InterDataNodeParameter]]] = {}
        for k, v in temp.items():
            temp2[k] = {CurrentPeriod: set(y.currentPeriod for y in v),
                        PreviousPeriod: set(y.previousPeriod for y in v)}

        return temp2

    def LoadAocStepConfiguration(self, year: int, month: int) -> list[AocConfiguration]:

        temp = {}
        for x in self.LoadParameter(AocConfiguration, year, month):
            key = (x.AocType, x.Novelty)
            temp.setdefault(key, []).append(x)

        result = []
        for k, v in temp.items():
            result.append(max(v, key=lambda y: y.Year * 100 + y.Month))

        return result

    def LoadAocStepConfigurationAsDictionary(self, year: int, month: int) -> dict[AocStep, AocConfiguration]:
        return {AocStep(x.AocType, x.Novelty): x for x in self.LoadAocStepConfiguration(year, month)}


    # Import helper

    def GetArgsFromMain(self, PartitionType: IKeyedType, dataSet: IDataSet) -> ImportArgs:

        mainTab: pd.DataFrame = dataSet.Tables[Main]

        main = mainTab.iloc[0]  # mainTab.Rows.First()
        reportingNode = main["ReportingNode"]
        scenario = main["Scenario"] if 'Scenario' in mainTab.columns and main["Scenario"] else ""   # Convert nan to ''

        args: ImportArgs

        if PartitionType is PartitionByReportingNode:
            args = ImportArgs(reportingNode, 0, 0, Periodicity.Monthly, scenario, "")
            self.Update(PartitionByReportingNode,
                [PartitionByReportingNode(
                             Id=self.Partition.GetKeyForInstance(PartitionByReportingNode, args),
                             ReportingNode=reportingNode,
                             Scenario=scenario)])

        elif PartitionType is PartitionByReportingNodeAndPeriod:

            if list(mainTab.columns).count('Year') != 1:
                raise YearInMainNotFound
            if list(mainTab.columns).count('Month') != 1:
                raise MonthInMainNotFound

            args = ImportArgs(reportingNode, int(main["Year"]), int(main["Month"]), Periodicity.Monthly, scenario, "")

            self.Update(PartitionByReportingNodeAndPeriod,
              [PartitionByReportingNodeAndPeriod(
                    Id=self.Partition.GetKeyForInstance(PartitionByReportingNodeAndPeriod, args),
                    Year=args.Year,
                    Month=args.Month,
                    ReportingNode=reportingNode,
                    Scenario=scenario)])

        else:
            # ApplicationMessage.Log(Error.PartitionTypeNotFound, typeof(IPartition).Name)
            raise PartitionTypeNotFound

        # await DataSource.CommitAsync()
        return args

    ## Data Node Factory
    def DataNodeFactory(self, dataSet: IDataSet, tableName: str, args: ImportArgs):
        partition = [p for p in self.Query(PartitionByReportingNode)
                     if p.ReportingNode == args.ReportingNode and p.Scenario == ''][0]

        if not partition:
            raise ParsedPartitionNotFound

        table = dataSet.Tables[tableName]

        dataNodesImported = remove_duplicates(table["DataNode"])
        dataNodesDefined = [x for x in self.Query(GroupOfContract) if x.SystemName in dataNodesImported]
        dataNodeStatesDefined = [x.DataNode for x in self.Query(DataNodeState)]
        dataNodeParametersDefined = [x.DataNode for x in self.Query(SingleDataNodeParameter)]

        dataNodeStatesUndefined = remove_duplicates([x for x in dataNodesImported if x and not x in dataNodeStatesDefined])
        dataNodeSingleParametersUndefined = remove_duplicates(
            [x for x in dataNodesImported if (x and
                not x in dataNodeParametersDefined and
                isinstance([y for y in dataNodesDefined if
                y.SystemName == x][0],
                GroupOfInsuranceContract))])

        self.Update3(
            [DataNodeState(DataNode=x,
                           Year=args.Year,
                           Month=DefaultDataNodeActivationMonth,
                           State=State.Active,
                           Partition=partition.Id,
                           Id=uuid.uuid4(),
                           Scenario=''
                           ) for x in dataNodeStatesUndefined]
        )

        self.Update3(
            [SingleDataNodeParameter(
                DataNode=x, Year=args.Year,
                Month=DefaultDataNodeActivationMonth,
                PremiumAllocation=DefaultPremiumExperienceAdjustmentFactor,
                Partition=partition.Id,
                Id=uuid.uuid4(),
                Scenario=''
            ) for x in dataNodeSingleParametersUndefined]
        )

    def get_ifrsvars(self, add_goc_attrs=False):

        data = []
        for v in self.Query(IfrsVariable):
            data.append({
                'Partition': v.Partition,
                'DataNode': v.DataNode,
                'AocType': v.AocType,
                'Novelty': v.Novelty,
                'AmountType': v.AmountType,
                'AccidentYear': v.AccidentYear,
                'EstimateType': v.EstimateType,
                'EconomicBasis': v.EconomicBasis,
                'Value': v.Value
            })
        vars = pd.DataFrame.from_records(data)

        data = []
        for v in self.Query(PartitionByReportingNodeAndPeriod):
            data.append({
                'Partition': v.Id,
                'ReportingNode': v.ReportingNode,
                'Year': v.Year,
                'Month': v.Month
            })
        part = pd.DataFrame.from_records(data, index='Partition')

        df = vars.join(part, on='Partition')
        df = df.loc[:, df.columns != 'Partition']

        if add_goc_attrs:
            return self._add_goc_attrs(df)
        else:
            return df

    def _add_goc_attrs(self, data: pd.DataFrame):

        goc = self._query2df(self.Query(GroupOfContract))   # super class
        goc.rename(columns={'SystemName': 'DataNode'}, inplace=True)
        goc.set_index('DataNode', inplace=True)

        df = pd.merge(data, goc, left_on='DataNode', right_index=True, how='left', sort=False)
        return df.loc[:, df.columns != 'Partition']

